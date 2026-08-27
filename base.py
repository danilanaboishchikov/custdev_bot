import sqlite3
import random
import string
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
PATH = str(BASE_DIR) + '/'
DB_NAME = str(BASE_DIR / 'db.db')

def initialize_database():
    (BASE_DIR / 'excel').mkdir(exist_ok=True)
    (BASE_DIR / 'content').mkdir(exist_ok=True)
    (BASE_DIR / 'files').mkdir(exist_ok=True)

    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS rates (id INTEGER, from_id INTEGER, text TEXT, rate INTEGER)')
        cursor.execute('CREATE TABLE IF NOT EXISTS tasks (task_id TEXT, creator_id INTEGER, name TEXT, about TEXT, target_people TEXT, comment TEXT, price INTEGER, worker TEXT, category TEXT, status TEXT, regs TEXT, report TEXT)')
        cursor.execute('CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY, name TEXT, username TEXT, type TEXT, money INTEGER, bufer_money INTEGER, all_money INTEGER, tasks_cnt INTEGER, rates REAL, rates_cnt INTEGER, sum_rates INTEGER, info TEXT, category TEXT, taken_tasks TEXT)')


initialize_database()

def get_conn():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

# Функции для таблицы rates
def add_rate(id, from_id, text, rate):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO rates (id, from_id, text, rate) VALUES (?, ?, ?, ?)",
                       (id, from_id, text, rate))

def get_rates_by_user_id(user_id):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM rates WHERE id = ?", (user_id,))
        return [dict(row) for row in cursor.fetchall()]

def get_rates_by_from_id(from_id):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM rates WHERE from_id = ?", (from_id,))
        return [dict(row) for row in cursor.fetchall()]

# Функции для таблицы tasks
def add_task(task_id, creator_id, target_people, comment, price, category, name, about):
    with get_conn() as conn:
        cursor = conn.cursor()
        category_str = '|'.join(category).strip()
        cursor.execute("INSERT INTO tasks (task_id, creator_id, target_people, comment, price, worker, category, status, name, about, regs, report) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                       (task_id, creator_id, target_people, comment, price, '-', category_str, 'free', name, about, '', ''))

def add_reg_to_task(task_id, reg):
    # Открываем соединение с базой данных
    with get_conn() as conn:
        cursor = conn.cursor()

        # Получаем текущее значение поля 'regs' для указанного task_id
        cursor.execute("SELECT regs FROM tasks WHERE task_id = ?", (task_id,))
        result = cursor.fetchone()

        if result:
            current_regs = result[0] or ""  # Если поле 'regs' пустое, задаем его как пустую строку

            # Если поле 'regs' не пустое, добавляем новый reg через '|'
            if current_regs != '':
                new_regs = f"{current_regs}|{reg}"
            else:
                new_regs = reg

            # Обновляем поле 'regs' для указанного task_id
            cursor.execute("UPDATE tasks SET regs = ? WHERE task_id = ?", (new_regs, task_id))
            conn.commit()  # Сохраняем изменения
        else:
            print(f"Задание с task_id {task_id} не найдено.")


def add_task_to_user(user_id, task_id):
    # Открываем соединение с базой данных
    with get_conn() as conn:
        cursor = conn.cursor()

        # Получаем текущее значение поля 'regs' для указанного task_id
        cursor.execute("SELECT taken_tasks FROM users WHERE id = ?", (user_id,))
        result = cursor.fetchone()

        if result:
            current_tasks = result[0] or ""  # Если поле 'regs' пустое, задаем его как пустую строку

            # Если поле 'regs' не пустое, добавляем новый reg через '|'
            if current_tasks != '':
                new_tasks = f"{current_tasks}|{task_id}"
            else:
                new_tasks = task_id

            # Обновляем поле 'regs' для указанного task_id
            cursor.execute("UPDATE users SET taken_tasks = ? WHERE id = ?", (new_tasks, user_id))
            conn.commit()  # Сохраняем изменения
        else:
            print(f"Задание с task_id {task_id} не найдено.")


def remove_task_to_user(user_id, task_id):
    # Открываем соединение с базой данных
    with get_conn() as conn:
        cursor = conn.cursor()

        # Получаем текущее значение поля 'regs' для указанного task_id
        cursor.execute("SELECT taken_tasks FROM users WHERE id = ?", (user_id,))
        result = cursor.fetchone()

        if result:
            current_tasks = result[0] or ""  # Если поле 'regs' пустое, задаем его как пустую строку

            new_tasks = current_tasks.replace(task_id, '').replace('||', '|')
            new_tasks = new_tasks.lstrip('|').rstrip('|')

            # Обновляем поле 'regs' для указанного task_id
            cursor.execute("UPDATE users SET taken_tasks = ? WHERE id = ?", (new_tasks, user_id))
            conn.commit()  # Сохраняем изменения
        else:
            print(f"Задание с task_id {task_id} не найдено.")


def set_report(task_id, report):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE tasks SET report = ? WHERE task_id = ?", (report, task_id))

def delete_task(task_id):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM tasks WHERE task_id = ?", (task_id,))

def get_task(task_id):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tasks WHERE task_id = ?", (task_id,))
        row = cursor.fetchone()
        return dict(row) if row else None


def get_user_tasks(user_id):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tasks WHERE creator_id = ?", (user_id,))
        rows = cursor.fetchall()
        tasks = [dict(row) for row in rows]
        return tasks


def get_all_tasks():
    with get_conn() as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM tasks")
        rows = cursor.fetchall()

        # Преобразуем результат в список словарей
        tasks = [dict(row) for row in rows]

        return tasks

def update_task_status(task_id, status, worker=None):
    with get_conn() as conn:
        cursor = conn.cursor()
        if worker:
            cursor.execute("UPDATE tasks SET status = ?, worker = ? WHERE task_id = ?",
                           (status, worker, task_id))
        else:
            cursor.execute("UPDATE tasks SET status = ? WHERE task_id = ?",
                           (status, task_id))

def get_user(user_id):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (int(user_id),))
        row = cursor.fetchone()
        return dict(row) if row else None

# Функции для таблицы users
def add_user(id, name, username, type, info, category):
    with get_conn() as conn:
        cursor = conn.cursor()
        if category != '-':
            category_str = '|'.join(category).strip()
        else:
            category_str = '-'
        cursor.execute("INSERT INTO users (id, name, username, type, money, bufer_money, all_money, tasks_cnt, rates, rates_cnt, sum_rates, info, category, taken_tasks) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                       (id, name, username, type, 0, 0, 0, 0, -1, 0, 0, info, category_str, ''))

def update_user_money(user_id, money_change=None, bufer_money_change=None):
    with get_conn() as conn:
        cursor = conn.cursor()
        if money_change:
            value = money_change.replace('!', '')
            if '+' in value:
                value = int(money_change.replace('+', ''))
                cursor.execute("UPDATE users SET money = money + ?, all_money = all_money + ? WHERE id = ?",
                               (value, value if '!' in money_change else 0, user_id))
            else:
                value = int(money_change.replace('-', ''))
                cursor.execute("UPDATE users SET money = money - ? WHERE id = ?",
                               (value, user_id))

        if bufer_money_change:
            value = bufer_money_change.replace('!', '')
            if '+' in value:
                value = int(bufer_money_change.replace('+', ''))
                cursor.execute("UPDATE users SET bufer_money = bufer_money + ? WHERE id = ?",
                               (value, user_id))
            else:
                value = int(bufer_money_change.replace('-', ''))
                cursor.execute("UPDATE users SET bufer_money = bufer_money - ? WHERE id = ?",
                               (value, user_id))

def update_user_info(user_id, info):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET info = ? WHERE id = ?", (info, user_id))

def update_user_category(user_id, category):
    with get_conn() as conn:
        cursor = conn.cursor()
        category_str = '|'.join(category).strip()
        cursor.execute("UPDATE users SET category = ? WHERE id = ?", (category_str, user_id))

def update_user_rate(user_id, new_rate):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT rates, rates_cnt, sum_rates FROM users WHERE id = ?", (user_id,))
        row = cursor.fetchone()
        if row:
            rates, rates_cnt, sum_rates = row['rates'], row['rates_cnt'], row['sum_rates']
            new_rates_cnt = int(rates_cnt) + 1
            new_sum_rates = int(sum_rates) + int(new_rate)
            new_rates = new_sum_rates / new_rates_cnt
            cursor.execute("UPDATE users SET rates = ?, rates_cnt = ?, sum_rates = ? WHERE id = ?",
                           (new_rates, new_rates_cnt, new_sum_rates, user_id))


def update_user_task_cnt(user_id):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT tasks_cnt FROM users WHERE id = ?", (user_id,))
        cur = cursor.fetchone()[0]
        cursor.execute("UPDATE users SET tasks_cnt = ? WHERE id = ?",
                           (str(int(cur+1)), user_id))


def generate_task_id():
    with get_conn() as conn:
        cursor = conn.cursor()
        while True:
            task_id = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            cursor.execute("SELECT task_id FROM tasks WHERE task_id = ?", (task_id,))
            if not cursor.fetchone():
                return task_id

def is_user_registered(user_id):
    with get_conn() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM users WHERE id = ?", (user_id,))
        return cursor.fetchone() is not None


import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side


def export_to_excel():
    def get_db():
        conn = sqlite3.connect(DB_NAME)
        return conn
    conn = get_db()
 # Извлечение данных из таблиц
    rates_df = pd.read_sql_query("SELECT * FROM rates", conn)
    tasks_df = pd.read_sql_query("SELECT * FROM tasks", conn)
    users_df = pd.read_sql_query("SELECT * FROM users", conn)
    # Создание Excel файла
    workbook = Workbook()
    # Функция для добавления DataFrame в Excel с форматированием
    def add_dataframe_to_excel(sheet_name, df):
        sheet = workbook.create_sheet(title=sheet_name)
        # Добавление заголовков
        for col in df.columns:
            cell = sheet.cell(row=1, column=df.columns.get_loc(col) + 1, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Желтый фон
        # Добавление данных
        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row):
                cell = sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
        # Установка ширины столбцов
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    # Добавление всех DataFrame в Excel
    add_dataframe_to_excel('rates', rates_df)
    add_dataframe_to_excel('tasks', tasks_df)
    add_dataframe_to_excel('users', users_df)
    # Удаляем лист по умолчанию, если он был создан
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    # Сохраняем файл
    workbook.save(PATH + 'output.xlsx')
    conn.close()
    return PATH + 'output.xlsx'

