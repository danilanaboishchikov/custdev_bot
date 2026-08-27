import openpyxl
from openpyxl.styles import Font
import os

from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent
PATH = str(BASE_DIR / 'excel') + '/'
Path(PATH).mkdir(parents=True, exist_ok=True)

def write_to_excel(message, data_dict, filename, write_user_info=False):

    '''
    Функция для универсальной записи данных в excel
    :param message: сообщение пользователя, отправившего данные
    :param data_dict: словарь с данными, {заголовок: данные}
    :param filename: имя файла для записи
    :param write_user_info: True or False, True, если нужно записать информацию о пользователе из message
    :return:
    '''

    # Проверяем, существует ли файл
    file_exists = os.path.isfile(filename)

    # Если файл не существует, создаем его
    if not file_exists:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Создаем заголовки
        headers = []
        if write_user_info:
            headers.extend(['ID пользователя', 'Имя', 'Username'])
        headers.extend(data_dict.keys())

        # Записываем заголовки в первую строку
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)  # Делаем текст жирным

        # Записываем данные пользователя и словаря
        row_data = []
        if write_user_info:
            row_data.extend([message.from_user.id, message.from_user.first_name, message.from_user.username])
        row_data.extend(data_dict.values())

        sheet.append(row_data)  # Добавляем данные на следующую строку

        # Сохраняем файл
        workbook.save(filename)
    else:
        # Если файл существует, открываем его и добавляем данные
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Записываем данные пользователя и словаря
        row_data = []
        if write_user_info:
            row_data.extend([message.from_user.id, message.from_user.first_name, message.from_user.username])
        row_data.extend(data_dict.values())

        sheet.append(row_data)  # Добавляем данные на следующую строку

        # Сохраняем файл
        workbook.save(filename)